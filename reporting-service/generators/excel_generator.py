"""
Excel Report Generator
Generates attendance reports in Excel format
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Generates attendance reports in Excel format"""

    def __init__(self, output_dir='reports/excel'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        # Define styles
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font = Font(color='FFFFFF', bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.present_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        self.absent_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    def generate_student_report(self, student_data, attendance_records, course_data):
        """
        Generate attendance report for a single student

        Args:
            student_data: Student information (id, name, department)
            attendance_records: List of attendance records
            course_data: Course information

        Returns:
            str: Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Student Attendance'

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'Student Attendance Report'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Student information
        row = 3
        ws[f'A{row}'] = 'Student ID:'
        ws[f'B{row}'] = student_data.get('student_id', 'N/A')
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Student Name:'
        ws[f'B{row}'] = student_data.get('name', 'N/A')
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Department:'
        ws[f'B{row}'] = student_data.get('department', 'N/A')

        row += 1
        ws[f'A{row}'] = 'Course:'
        ws[f'B{row}'] = course_data.get('name', 'N/A')

        # Statistics
        total_lectures = len(attendance_records)
        present_count = sum(1 for r in attendance_records if r.get('status') == 'present')
        absent_count = total_lectures - present_count
        attendance_percentage = (present_count / total_lectures * 100) if total_lectures > 0 else 0

        row += 1
        ws[f'A{row}'] = 'Total Lectures:'
        ws[f'B{row}'] = total_lectures

        row += 1
        ws[f'A{row}'] = 'Present:'
        ws[f'B{row}'] = present_count
        ws[f'B{row}'].fill = self.present_fill

        row += 1
        ws[f'A{row}'] = 'Absent:'
        ws[f'B{row}'] = absent_count
        ws[f'B{row}'].fill = self.absent_fill

        row += 1
        ws[f'A{row}'] = 'Attendance %:'
        ws[f'B{row}'] = f'{attendance_percentage:.2f}%'
        ws[f'B{row}'].font = Font(bold=True, size=12)

        # Attendance records table
        row += 3
        headers = ['Date', 'Lecture ID', 'Status', 'Marked At', 'Notes']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        # Data rows
        for record in attendance_records:
            row += 1
            ws.cell(row=row, column=1, value=record.get('date', '')).border = self.border
            ws.cell(row=row, column=2, value=record.get('lecture_id', '')).border = self.border

            status_cell = ws.cell(row=row, column=3, value=record.get('status', '').upper())
            status_cell.border = self.border
            if record.get('status') == 'present':
                status_cell.fill = self.present_fill
            else:
                status_cell.fill = self.absent_fill

            ws.cell(row=row, column=4, value=record.get('marked_at', '')).border = self.border
            ws.cell(row=row, column=5, value=record.get('notes', '')).border = self.border

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30

        # Save file
        filename = f"student_{student_data.get('student_id', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)

        return filepath

    def generate_course_report(self, course_data, lectures_data, students_data, attendance_matrix):
        """
        Generate attendance report for a course

        Args:
            course_data: Course information
            lectures_data: List of lectures
            students_data: List of students
            attendance_matrix: Dict mapping (student_id, lecture_id) -> status

        Returns:
            str: Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Course Attendance'

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Attendance Report - {course_data.get('name', 'Course')}"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Course information
        row = 3
        ws[f'A{row}'] = 'Course Code:'
        ws[f'B{row}'] = course_data.get('course_code', 'N/A')
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = 'Instructor:'
        ws[f'B{row}'] = course_data.get('instructor', 'N/A')

        row += 1
        ws[f'A{row}'] = 'Total Students:'
        ws[f'B{row}'] = len(students_data)

        row += 1
        ws[f'A{row}'] = 'Total Lectures:'
        ws[f'B{row}'] = len(lectures_data)

        # Attendance matrix
        row += 3

        # Headers
        ws.cell(row=row, column=1, value='Student ID').fill = self.header_fill
        ws.cell(row=row, column=1).font = self.header_font
        ws.cell(row=row, column=1).border = self.border

        ws.cell(row=row, column=2, value='Student Name').fill = self.header_fill
        ws.cell(row=row, column=2).font = self.header_font
        ws.cell(row=row, column=2).border = self.border

        # Lecture dates as columns
        for idx, lecture in enumerate(lectures_data, start=3):
            cell = ws.cell(row=row, column=idx, value=lecture.get('date', f"L{idx-2}"))
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', text_rotation=90)
            cell.border = self.border

        # Present/Absent/Percentage columns
        present_col = len(lectures_data) + 3
        absent_col = present_col + 1
        percentage_col = absent_col + 1

        ws.cell(row=row, column=present_col, value='Present').fill = self.header_fill
        ws.cell(row=row, column=present_col).font = self.header_font
        ws.cell(row=row, column=present_col).border = self.border

        ws.cell(row=row, column=absent_col, value='Absent').fill = self.header_fill
        ws.cell(row=row, column=absent_col).font = self.header_font
        ws.cell(row=row, column=absent_col).border = self.border

        ws.cell(row=row, column=percentage_col, value='%').fill = self.header_fill
        ws.cell(row=row, column=percentage_col).font = self.header_font
        ws.cell(row=row, column=percentage_col).border = self.border

        # Student rows
        for student in students_data:
            row += 1
            student_id = student.get('student_id', '')

            ws.cell(row=row, column=1, value=student_id).border = self.border
            ws.cell(row=row, column=2, value=student.get('name', '')).border = self.border

            present_count = 0
            absent_count = 0

            # Attendance for each lecture
            for idx, lecture in enumerate(lectures_data, start=3):
                lecture_id = lecture.get('lecture_id', '')
                status = attendance_matrix.get((student_id, lecture_id), 'absent')

                cell = ws.cell(row=row, column=idx, value='P' if status == 'present' else 'A')
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')

                if status == 'present':
                    cell.fill = self.present_fill
                    present_count += 1
                else:
                    cell.fill = self.absent_fill
                    absent_count += 1

            # Statistics
            total = len(lectures_data)
            percentage = (present_count / total * 100) if total > 0 else 0

            ws.cell(row=row, column=present_col, value=present_count).border = self.border
            ws.cell(row=row, column=absent_col, value=absent_count).border = self.border

            percentage_cell = ws.cell(row=row, column=percentage_col, value=f'{percentage:.1f}%')
            percentage_cell.border = self.border
            if percentage < 50:
                percentage_cell.fill = self.absent_fill
            elif percentage < 75:
                percentage_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                percentage_cell.fill = self.present_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        for idx in range(3, len(lectures_data) + 3):
            ws.column_dimensions[get_column_letter(idx)].width = 5
        ws.column_dimensions[get_column_letter(present_col)].width = 10
        ws.column_dimensions[get_column_letter(absent_col)].width = 10
        ws.column_dimensions[get_column_letter(percentage_col)].width = 10

        # Save file
        filename = f"course_{course_data.get('course_code', 'unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)

        return filepath

    def generate_department_report(self, department, courses_stats):
        """
        Generate department-wide attendance summary

        Args:
            department: Department name
            courses_stats: List of course statistics

        Returns:
            str: Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Department Summary'

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f'Department Attendance Summary - {department}'
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        row = 3
        headers = ['Course Code', 'Course Name', 'Total Students', 'Avg Attendance %', 'At Risk Students', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border

        # Data rows
        for course in courses_stats:
            row += 1
            ws.cell(row=row, column=1, value=course.get('course_code', '')).border = self.border
            ws.cell(row=row, column=2, value=course.get('course_name', '')).border = self.border
            ws.cell(row=row, column=3, value=course.get('total_students', 0)).border = self.border

            avg_attendance = course.get('avg_attendance', 0)
            avg_cell = ws.cell(row=row, column=4, value=f'{avg_attendance:.2f}%')
            avg_cell.border = self.border
            if avg_attendance < 50:
                avg_cell.fill = self.absent_fill
            elif avg_attendance < 75:
                avg_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                avg_cell.fill = self.present_fill

            ws.cell(row=row, column=5, value=course.get('at_risk_students', 0)).border = self.border

            status = 'Good' if avg_attendance >= 75 else 'Warning' if avg_attendance >= 50 else 'Critical'
            status_cell = ws.cell(row=row, column=6, value=status)
            status_cell.border = self.border
            if status == 'Critical':
                status_cell.fill = self.absent_fill
            elif status == 'Warning':
                status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            else:
                status_cell.fill = self.present_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12

        # Save file
        filename = f"department_{department}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)

        return filepath

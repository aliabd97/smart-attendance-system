"""
Create sample Excel file for testing Excel Adapter Pattern
Generates legacy-format Excel file with student data
"""

import openpyxl
from datetime import datetime, timedelta
import random

# Sample Iraqi student names
FIRST_NAMES = [
    'Ahmed', 'Mohammed', 'Ali', 'Hussein', 'Omar', 'Khalid', 'Hassan', 'Mustafa',
    'Fatima', 'Zainab', 'Mariam', 'Sarah', 'Noor', 'Hiba', 'Zahraa', 'Rana'
]

LAST_NAMES = [
    'Abdullah', 'Al-Saadi', 'Al-Maliki', 'Al-Baghdadi', 'Al-Basri', 'Al-Najafi',
    'Mohammed', 'Hussein', 'Hassan', 'Ali', 'Al-Mosawi', 'Al-Karbalai'
]

DEPARTMENTS = {
    'CS': 'Computer Science',
    'IT': 'Information Technology',
    'SE': 'Software Engineering',
    'CE': 'Computer Engineering',
    'IS': 'Information Systems'
}

def generate_student_id(index):
    """Generate student ID"""
    year = 2021
    return f"{year}{index:04d}"

def generate_email(name, student_id):
    """Generate email address"""
    name_part = name.lower().replace(' ', '.')
    return f"{name_part}@university.edu.iq"

def generate_phone():
    """Generate Iraqi phone number"""
    prefixes = ['0770', '0771', '0772', '0780', '0781', '0782', '0790', '0791']
    return f"{random.choice(prefixes)}{random.randint(1000000, 9999999)}"

def generate_registration_date(index):
    """Generate registration date"""
    base_date = datetime(2021, 9, 1)  # September 1, 2021
    days_offset = index % 30  # Spread over 30 days
    return (base_date + timedelta(days=days_offset)).strftime('%d/%m/%Y')

def create_sample_excel(filename='sample_students.xlsx', num_students=50):
    """
    Create sample Excel file with student data

    Args:
        filename: Output filename
        num_students: Number of students to generate
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Headers (using legacy field names)
    headers = [
        'Student_No',
        'Full_Name',
        'Email_Address',
        'Dept',
        'Year',
        'Mobile',
        'Status',
        'Reg_Date'
    ]

    ws.append(headers)

    # Generate student data
    for i in range(1, num_students + 1):
        first_name = random.choice(FIRST_NAMES)
        last_name = random.choice(LAST_NAMES)
        full_name = f"{first_name} {last_name}"

        student_id = generate_student_id(i)
        email = generate_email(full_name, student_id)
        dept_code = random.choice(list(DEPARTMENTS.keys()))
        level = random.randint(1, 4)
        phone = generate_phone()
        status = 'Active' if random.random() > 0.1 else 'Inactive'  # 90% active
        reg_date = generate_registration_date(i)

        row = [
            student_id,
            full_name,
            email,
            dept_code,
            level,
            phone,
            status,
            reg_date
        ]

        ws.append(row)

    # Format cells
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Bold headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Save file
    wb.save(filename)
    print(f"✅ Created sample Excel file: {filename}")
    print(f"📊 Generated {num_students} student records")
    print(f"\nSample data preview:")
    print(f"Student ID: {ws['A2'].value}")
    print(f"Name: {ws['B2'].value}")
    print(f"Email: {ws['C2'].value}")
    print(f"Department: {ws['D2'].value}")
    print(f"Level: {ws['E2'].value}")

if __name__ == '__main__':
    print("=" * 50)
    print("📝 Sample Excel File Generator")
    print("=" * 50)
    create_sample_excel('sample_students.xlsx', num_students=50)
    print("=" * 50)
